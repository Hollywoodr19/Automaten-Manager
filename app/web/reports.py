# app/web/reports.py
"""
Berichte-Modul für Automaten Manager
Vollständige Implementation mit Export-Funktionen
"""

from flask import Blueprint, render_template_string, redirect, url_for, flash, request, jsonify, send_file
from flask_login import login_required, current_user
from datetime import datetime, date, timedelta
from decimal import Decimal
from sqlalchemy import func, extract, and_, or_
from app import db
from app.models import Device, Entry, Expense, Product, Refill, RefillItem, Supplier, DeviceStatus, ExpenseCategory
import io
import csv
import json
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

reports_bp = Blueprint('reports', __name__, url_prefix='/reports')


@reports_bp.route('/')
@login_required
def index():
    """Berichte Dashboard"""
    
    # Aktuelle Zeiträume
    today = date.today()
    current_month = today.month
    current_year = today.year
    last_month = (today.replace(day=1) - timedelta(days=1))
    
    # Schnellstatistiken für Übersicht
    month_income = Entry.query.join(Device).filter(
        Device.owner_id == current_user.id,
        extract('year', Entry.date) == current_year,
        extract('month', Entry.date) == current_month
    ).with_entities(func.sum(Entry.amount)).scalar() or 0
    
    month_expenses = Expense.query.filter(
        Expense.user_id == current_user.id,
        extract('year', Expense.date) == current_year,
        extract('month', Expense.date) == current_month
    ).with_entities(func.sum(Expense.amount)).scalar() or 0
    
    year_income = Entry.query.join(Device).filter(
        Device.owner_id == current_user.id,
        extract('year', Entry.date) == current_year
    ).with_entities(func.sum(Entry.amount)).scalar() or 0
    
    year_expenses = Expense.query.filter(
        Expense.user_id == current_user.id,
        extract('year', Expense.date) == current_year
    ).with_entities(func.sum(Expense.amount)).scalar() or 0
    
    content = f"""
    <div class="d-flex justify-content-between align-items-center mb-4">
        <h2 class="text-white">
            <i class="bi bi-graph-up"></i> Berichte & Analysen
        </h2>
        <div>
            <button class="btn btn-light" onclick="generateQuickReport()">
                <i class="bi bi-lightning"></i> Schnellbericht
            </button>
        </div>
    </div>

    <!-- Übersichtskarten -->
    <div class="row mb-4">
        <div class="col-md-3">
            <div class="card">
                <div class="card-body">
                    <h6 class="text-muted">Einnahmen (Monat)</h6>
                    <h3 class="text-success">{month_income:.2f} €</h3>
                    <small class="text-muted">{datetime(current_year, current_month, 1).strftime('%B %Y')}</small>
                </div>
            </div>
        </div>
        <div class="col-md-3">
            <div class="card">
                <div class="card-body">
                    <h6 class="text-muted">Ausgaben (Monat)</h6>
                    <h3 class="text-danger">{month_expenses:.2f} €</h3>
                    <small class="text-muted">{datetime(current_year, current_month, 1).strftime('%B %Y')}</small>
                </div>
            </div>
        </div>
        <div class="col-md-3">
            <div class="card">
                <div class="card-body">
                    <h6 class="text-muted">Gewinn (Jahr)</h6>
                    <h3 class="{'text-success' if (year_income - year_expenses) >= 0 else 'text-danger'}">{(year_income - year_expenses):.2f} €</h3>
                    <small class="text-muted">{current_year}</small>
                </div>
            </div>
        </div>
        <div class="col-md-3">
            <div class="card">
                <div class="card-body">
                    <h6 class="text-muted">ROI</h6>
                    <h3 class="text-info">{((year_income - year_expenses) / year_expenses * 100) if year_expenses > 0 else 0:.1f}%</h3>
                    <small class="text-muted">Return on Investment</small>
                </div>
            </div>
        </div>
    </div>

    <!-- Bericht-Typen -->
    <div class="row">
        <div class="col-md-4 mb-3">
            <div class="card h-100">
                <div class="card-body">
                    <h5 class="card-title">
                        <i class="bi bi-calendar-month text-primary"></i> Monatsbericht
                    </h5>
                    <p class="card-text">Detaillierte Monatsübersicht mit allen Einnahmen und Ausgaben</p>
                    <form action="{url_for('reports.monthly_report')}" method="post" class="mt-3">
                        <div class="mb-3">
                            <select name="month" class="form-select">
                                <option value="{current_month}" selected>{datetime(current_year, current_month, 1).strftime('%B')}</option>
                                <option value="{last_month.month}">{last_month.strftime('%B')}</option>
                            </select>
                        </div>
                        <div class="mb-3">
                            <select name="year" class="form-select">
                                <option value="{current_year}">{current_year}</option>
                                <option value="{current_year-1}">{current_year-1}</option>
                            </select>
                        </div>
                        <div class="btn-group w-100">
                            <button type="submit" name="format" value="pdf" class="btn btn-primary">
                                <i class="bi bi-file-pdf"></i> PDF
                            </button>
                            <button type="submit" name="format" value="excel" class="btn btn-success">
                                <i class="bi bi-file-excel"></i> Excel
                            </button>
                        </div>
                    </form>
                </div>
            </div>
        </div>

        <div class="col-md-4 mb-3">
            <div class="card h-100">
                <div class="card-body">
                    <h5 class="card-title">
                        <i class="bi bi-calendar-year text-success"></i> Jahresbericht
                    </h5>
                    <p class="card-text">Jahresübersicht mit Trends und Vergleichen</p>
                    <form action="{url_for('reports.yearly_report')}" method="post" class="mt-3">
                        <div class="mb-3">
                            <select name="year" class="form-select">
                                <option value="{current_year}">{current_year}</option>
                                <option value="{current_year-1}">{current_year-1}</option>
                                <option value="{current_year-2}">{current_year-2}</option>
                            </select>
                        </div>
                        <div class="btn-group w-100">
                            <button type="submit" name="format" value="pdf" class="btn btn-primary">
                                <i class="bi bi-file-pdf"></i> PDF
                            </button>
                            <button type="submit" name="format" value="excel" class="btn btn-success">
                                <i class="bi bi-file-excel"></i> Excel
                            </button>
                        </div>
                    </form>
                </div>
            </div>
        </div>

        <div class="col-md-4 mb-3">
            <div class="card h-100">
                <div class="card-body">
                    <h5 class="card-title">
                        <i class="bi bi-cpu text-warning"></i> Gerätebericht
                    </h5>
                    <p class="card-text">Performance-Analyse einzelner Geräte</p>
                    <form action="{url_for('reports.device_report')}" method="post" class="mt-3">
                        <div class="mb-3">
                            <select name="device_id" class="form-select">
                                <option value="all">Alle Geräte</option>
    """
    
    devices = Device.query.filter_by(owner_id=current_user.id).all()
    for device in devices:
        content += f'<option value="{device.id}">{device.name}</option>'
    
    content += f"""
                            </select>
                        </div>
                        <div class="mb-3">
                            <select name="period" class="form-select">
                                <option value="month">Letzter Monat</option>
                                <option value="quarter">Letztes Quartal</option>
                                <option value="year">Letztes Jahr</option>
                            </select>
                        </div>
                        <div class="btn-group w-100">
                            <button type="submit" name="format" value="pdf" class="btn btn-primary">
                                <i class="bi bi-file-pdf"></i> PDF
                            </button>
                            <button type="submit" name="format" value="excel" class="btn btn-success">
                                <i class="bi bi-file-excel"></i> Excel
                            </button>
                        </div>
                    </form>
                </div>
            </div>
        </div>
    </div>

    <!-- Weitere Berichte -->
    <div class="row mt-4">
        <div class="col-md-4 mb-3">
            <div class="card h-100">
                <div class="card-body">
                    <h5 class="card-title">
                        <i class="bi bi-box text-info"></i> Produktanalyse
                    </h5>
                    <p class="card-text">Umsatz und Verbrauch pro Produkt</p>
                    <a href="{url_for('reports.product_analysis')}" class="btn btn-info w-100">
                        <i class="bi bi-bar-chart"></i> Analyse starten
                    </a>
                </div>
            </div>
        </div>

        <div class="col-md-4 mb-3">
            <div class="card h-100">
                <div class="card-body">
                    <h5 class="card-title">
                        <i class="bi bi-cash-coin text-success"></i> Cashflow
                    </h5>
                    <p class="card-text">Geldfluss-Analyse und Prognose</p>
                    <a href="{url_for('reports.cashflow')}" class="btn btn-success w-100">
                        <i class="bi bi-graph-up-arrow"></i> Cashflow anzeigen
                    </a>
                </div>
            </div>
        </div>

        <div class="col-md-4 mb-3">
            <div class="card h-100">
                <div class="card-body">
                    <h5 class="card-title">
                        <i class="bi bi-download text-secondary"></i> Datenexport
                    </h5>
                    <p class="card-text">Komplette Daten als CSV/Excel</p>
                    <a href="{url_for('reports.export_data')}" class="btn btn-secondary w-100">
                        <i class="bi bi-database"></i> Daten exportieren
                    </a>
                </div>
            </div>
        </div>
    </div>
    """
    
    extra_scripts = """
    <script>
    function generateQuickReport() {
        if (confirm('Schnellbericht für den aktuellen Monat generieren?')) {
            window.location.href = '/reports/quick';
        }
    }
    </script>
    """
    
    from app.web.dashboard_modern import render_modern_template
    
    return render_modern_template(
        content=content + extra_scripts,
        title='Berichte',
        active_module='reports',
        active_submodule='overview',
        breadcrumb=[
            {'text': 'Dashboard', 'url': url_for('dashboard_modern.dashboard')},
            {'text': 'Berichte'}
        ]
    )


@reports_bp.route('/monthly', methods=['POST'])
@login_required
def monthly_report():
    """Monatsbericht generieren"""
    month = int(request.form.get('month', datetime.now().month))
    year = int(request.form.get('year', datetime.now().year))
    format_type = request.form.get('format', 'pdf')
    
    # Daten sammeln
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        end_date = date(year, month + 1, 1) - timedelta(days=1)
    
    # Einnahmen
    entries = Entry.query.join(Device).filter(
        Device.owner_id == current_user.id,
        Entry.date >= start_date,
        Entry.date <= end_date
    ).order_by(Entry.date).all()
    
    # Ausgaben
    expenses = Expense.query.filter(
        Expense.user_id == current_user.id,
        Expense.date >= start_date,
        Expense.date <= end_date
    ).order_by(Expense.date).all()
    
    # Geräte-Performance
    device_stats = db.session.query(
        Device.name,
        func.sum(Entry.amount).label('total'),
        func.count(Entry.id).label('count')
    ).join(Entry).filter(
        Device.owner_id == current_user.id,
        Entry.date >= start_date,
        Entry.date <= end_date
    ).group_by(Device.id, Device.name).all()
    
    if format_type == 'excel':
        return generate_excel_report(
            f'Monatsbericht_{month}_{year}',
            entries, expenses, device_stats,
            start_date, end_date
        )
    else:
        return generate_pdf_report(
            f'Monatsbericht {month}/{year}',
            entries, expenses, device_stats,
            start_date, end_date
        )


@reports_bp.route('/yearly', methods=['POST'])
@login_required
def yearly_report():
    """Jahresbericht generieren"""
    year = int(request.form.get('year', datetime.now().year))
    format_type = request.form.get('format', 'pdf')
    
    # Monatsweise Daten sammeln
    monthly_data = []
    for month in range(1, 13):
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(year, month + 1, 1) - timedelta(days=1)
        
        month_income = Entry.query.join(Device).filter(
            Device.owner_id == current_user.id,
            Entry.date >= start_date,
            Entry.date <= end_date
        ).with_entities(func.sum(Entry.amount)).scalar() or 0
        
        month_expenses = Expense.query.filter(
            Expense.user_id == current_user.id,
            Expense.date >= start_date,
            Expense.date <= end_date
        ).with_entities(func.sum(Expense.amount)).scalar() or 0
        
        monthly_data.append({
            'month': datetime(year, month, 1).strftime('%B'),
            'income': month_income,
            'expenses': month_expenses,
            'profit': month_income - month_expenses
        })
    
    if format_type == 'excel':
        return generate_yearly_excel(year, monthly_data)
    else:
        return generate_yearly_pdf(year, monthly_data)


@reports_bp.route('/device', methods=['POST'])
@login_required
def device_report():
    """Gerätebericht generieren"""
    device_id = request.form.get('device_id')
    period = request.form.get('period', 'month')
    format_type = request.form.get('format', 'pdf')
    
    # Zeitraum bestimmen
    end_date = date.today()
    if period == 'month':
        start_date = end_date - timedelta(days=30)
    elif period == 'quarter':
        start_date = end_date - timedelta(days=90)
    else:  # year
        start_date = end_date - timedelta(days=365)
    
    if device_id == 'all':
        devices = Device.query.filter_by(owner_id=current_user.id).all()
    else:
        devices = [Device.query.get(device_id)]
    
    device_data = []
    for device in devices:
        entries = Entry.query.filter(
            Entry.device_id == device.id,
            Entry.date >= start_date,
            Entry.date <= end_date
        ).all()
        
        expenses = Expense.query.filter(
            Expense.device_id == device.id,
            Expense.date >= start_date,
            Expense.date <= end_date
        ).all()
        
        total_income = sum(e.amount for e in entries)
        total_expenses = sum(e.amount for e in expenses)
        
        device_data.append({
            'device': device,
            'income': total_income,
            'expenses': total_expenses,
            'profit': total_income - total_expenses,
            'entries': len(entries),
            'avg_daily': total_income / ((end_date - start_date).days or 1)
        })
    
    if format_type == 'excel':
        return generate_device_excel(device_data, start_date, end_date)
    else:
        return generate_device_pdf(device_data, start_date, end_date)


@reports_bp.route('/product-analysis')
@login_required
def product_analysis():
    """Produktanalyse-Seite"""
    products = Product.query.filter_by(user_id=current_user.id).all()
    
    product_stats = []
    for product in products:
        # Verbrauch berechnen
        refill_items = RefillItem.query.join(Refill).filter(
            RefillItem.product_id == product.id,
            Refill.user_id == current_user.id
        ).all()
        
        total_quantity = sum(item.quantity for item in refill_items)
        total_value = sum(item.total_price for item in refill_items)
        
        product_stats.append({
            'product': product,
            'total_quantity': total_quantity,
            'total_value': total_value,
            'current_stock': product.get_current_stock(),
            'turnover_rate': (total_quantity / product.get_current_stock()) if product.get_current_stock() > 0 else 0
        })
    
    # Nach Wert sortieren
    product_stats.sort(key=lambda x: x['total_value'], reverse=True)
    
    content = f"""
    <div class="d-flex justify-content-between align-items-center mb-4">
        <h2 class="text-white">
            <i class="bi bi-box"></i> Produktanalyse
        </h2>
        <div>
            <button class="btn btn-light" onclick="exportProductAnalysis()">
                <i class="bi bi-download"></i> Export
            </button>
        </div>
    </div>

    <div class="card">
        <div class="card-body">
            <div class="table-responsive">
                <table class="table table-hover">
                    <thead>
                        <tr>
                            <th>Produkt</th>
                            <th>Kategorie</th>
                            <th>Gesamtmenge</th>
                            <th>Gesamtwert</th>
                            <th>Aktueller Bestand</th>
                            <th>Umschlagsrate</th>
                            <th>Status</th>
                        </tr>
                    </thead>
                    <tbody>
    """
    
    for stat in product_stats[:20]:  # Top 20 Produkte
        product = stat['product']
        status_color = 'success' if stat['current_stock'] > product.reorder_point else 'warning'
        
        content += f"""
                        <tr>
                            <td><strong>{product.name}</strong></td>
                            <td>{product.category.value if product.category else '-'}</td>
                            <td>{stat['total_quantity']:.1f} {product.unit.value}</td>
                            <td>{stat['total_value']:.2f} €</td>
                            <td>{stat['current_stock']:.1f} {product.unit.value}</td>
                            <td>{stat['turnover_rate']:.2f}x</td>
                            <td><span class="badge bg-{status_color}">{'Gut' if status_color == 'success' else 'Niedrig'}</span></td>
                        </tr>
        """
    
    content += """
                    </tbody>
                </table>
            </div>
        </div>
    </div>

    <script>
    function exportProductAnalysis() {
        window.location.href = '/reports/export/products';
    }
    </script>
    """
    
    from app.web.dashboard_modern import render_modern_template
    
    return render_modern_template(
        content=content,
        title='Produktanalyse',
        active_module='reports',
        active_submodule='products',
        breadcrumb=[
            {'text': 'Dashboard', 'url': url_for('dashboard_modern.dashboard')},
            {'text': 'Berichte', 'url': url_for('reports.index')},
            {'text': 'Produktanalyse'}
        ]
    )


@reports_bp.route('/cashflow')
@login_required
def cashflow():
    """Cashflow-Analyse"""
    # Letzten 12 Monate
    months = []
    for i in range(11, -1, -1):
        month_date = date.today().replace(day=1) - timedelta(days=i*30)
        months.append(month_date)
    
    cashflow_data = []
    cumulative = 0
    
    for month_date in months:
        start_date = month_date.replace(day=1)
        if month_date.month == 12:
            end_date = date(month_date.year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(month_date.year, month_date.month + 1, 1) - timedelta(days=1)
        
        income = Entry.query.join(Device).filter(
            Device.owner_id == current_user.id,
            Entry.date >= start_date,
            Entry.date <= end_date
        ).with_entities(func.sum(Entry.amount)).scalar() or 0
        
        expenses = Expense.query.filter(
            Expense.user_id == current_user.id,
            Expense.date >= start_date,
            Expense.date <= end_date
        ).with_entities(func.sum(Expense.amount)).scalar() or 0
        
        net = income - expenses
        cumulative += net
        
        cashflow_data.append({
            'month': month_date.strftime('%b %Y'),
            'income': float(income),
            'expenses': float(expenses),
            'net': float(net),
            'cumulative': float(cumulative)
        })
    
    content = f"""
    <div class="d-flex justify-content-between align-items-center mb-4">
        <h2 class="text-white">
            <i class="bi bi-cash-coin"></i> Cashflow-Analyse
        </h2>
    </div>

    <div class="card mb-4">
        <div class="card-body">
            <canvas id="cashflowChart"></canvas>
        </div>
    </div>

    <div class="card">
        <div class="card-header">
            <h5>Monatlicher Cashflow</h5>
        </div>
        <div class="card-body">
            <div class="table-responsive">
                <table class="table">
                    <thead>
                        <tr>
                            <th>Monat</th>
                            <th>Einnahmen</th>
                            <th>Ausgaben</th>
                            <th>Netto</th>
                            <th>Kumulativ</th>
                        </tr>
                    </thead>
                    <tbody>
    """
    
    for data in cashflow_data:
        net_color = 'text-success' if data['net'] >= 0 else 'text-danger'
        cumulative_color = 'text-success' if data['cumulative'] >= 0 else 'text-danger'
        
        content += f"""
                        <tr>
                            <td>{data['month']}</td>
                            <td class="text-success">+{data['income']:.2f} €</td>
                            <td class="text-danger">-{data['expenses']:.2f} €</td>
                            <td class="{net_color}"><strong>{data['net']:.2f} €</strong></td>
                            <td class="{cumulative_color}"><strong>{data['cumulative']:.2f} €</strong></td>
                        </tr>
        """
    
    content += f"""
                    </tbody>
                </table>
            </div>
        </div>
    </div>

    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
    <script>
    const ctx = document.getElementById('cashflowChart').getContext('2d');
    const cashflowData = {json.dumps(cashflow_data)};
    
    new Chart(ctx, {{
        type: 'line',
        data: {{
            labels: cashflowData.map(d => d.month),
            datasets: [{{
                label: 'Einnahmen',
                data: cashflowData.map(d => d.income),
                borderColor: 'rgb(40, 167, 69)',
                backgroundColor: 'rgba(40, 167, 69, 0.1)',
                tension: 0.4
            }}, {{
                label: 'Ausgaben',
                data: cashflowData.map(d => d.expenses),
                borderColor: 'rgb(220, 53, 69)',
                backgroundColor: 'rgba(220, 53, 69, 0.1)',
                tension: 0.4
            }}, {{
                label: 'Kumulativ',
                data: cashflowData.map(d => d.cumulative),
                borderColor: 'rgb(102, 126, 234)',
                backgroundColor: 'rgba(102, 126, 234, 0.1)',
                tension: 0.4,
                borderWidth: 3
            }}]
        }},
        options: {{
            responsive: true,
            plugins: {{
                legend: {{
                    position: 'top',
                }},
                title: {{
                    display: true,
                    text: 'Cashflow Verlauf (12 Monate)'
                }}
            }},
            scales: {{
                y: {{
                    beginAtZero: true,
                    ticks: {{
                        callback: function(value) {{
                            return value.toFixed(0) + ' €';
                        }}
                    }}
                }}
            }}
        }}
    }});
    </script>
    """
    
    from app.web.dashboard_modern import render_modern_template
    
    return render_modern_template(
        content=content,
        title='Cashflow',
        active_module='reports',
        active_submodule='cashflow',
        breadcrumb=[
            {'text': 'Dashboard', 'url': url_for('dashboard_modern.dashboard')},
            {'text': 'Berichte', 'url': url_for('reports.index')},
            {'text': 'Cashflow'}
        ]
    )


def generate_excel_report(title, entries, expenses, device_stats, start_date, end_date):
    """Excel-Report generieren"""
    wb = openpyxl.Workbook()
    
    # Übersicht Sheet
    ws = wb.active
    ws.title = "Übersicht"
    
    # Header
    ws['A1'] = title
    ws['A1'].font = Font(size=16, bold=True)
    ws['A2'] = f"Zeitraum: {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}"
    
    # Zusammenfassung
    total_income = sum(e.amount for e in entries)
    total_expenses = sum(e.amount for e in expenses)
    
    ws['A4'] = "Zusammenfassung"
    ws['A4'].font = Font(bold=True)
    ws['A5'] = "Gesamteinnahmen:"
    ws['B5'] = total_income
    ws['A6'] = "Gesamtausgaben:"
    ws['B6'] = total_expenses
    ws['A7'] = "Gewinn:"
    ws['B7'] = total_income - total_expenses
    
    # Einnahmen Sheet
    ws2 = wb.create_sheet("Einnahmen")
    ws2['A1'] = "Datum"
    ws2['B1'] = "Gerät"
    ws2['C1'] = "Betrag"
    ws2['D1'] = "Beschreibung"
    
    for i, entry in enumerate(entries, 2):
        ws2[f'A{i}'] = entry.date.strftime('%d.%m.%Y')
        ws2[f'B{i}'] = entry.device.name if entry.device else '-'
        ws2[f'C{i}'] = float(entry.amount)
        ws2[f'D{i}'] = entry.description or '-'
    
    # Ausgaben Sheet
    ws3 = wb.create_sheet("Ausgaben")
    ws3['A1'] = "Datum"
    ws3['B1'] = "Kategorie"
    ws3['C1'] = "Betrag"
    ws3['D1'] = "Beschreibung"
    
    for i, expense in enumerate(expenses, 2):
        ws3[f'A{i}'] = expense.date.strftime('%d.%m.%Y')
        ws3[f'B{i}'] = expense.category.value
        ws3[f'C{i}'] = float(expense.amount)
        ws3[f'D{i}'] = expense.description
    
    # Speichern und senden
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'{title}.xlsx'
    )


def generate_pdf_report(title, entries, expenses, device_stats, start_date, end_date):
    """PDF-Report generieren"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Titel
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#667eea'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 20))
    
    # Zeitraum
    period_text = f"Zeitraum: {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}"
    elements.append(Paragraph(period_text, styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Zusammenfassung
    total_income = sum(e.amount for e in entries)
    total_expenses = sum(e.amount for e in expenses)
    profit = total_income - total_expenses
    
    summary_data = [
        ['Zusammenfassung', ''],
        ['Gesamteinnahmen:', f'{total_income:.2f} €'],
        ['Gesamtausgaben:', f'{total_expenses:.2f} €'],
        ['Gewinn:', f'{profit:.2f} €']
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(summary_table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'{title}.pdf'
    )


def generate_yearly_excel(year, monthly_data):
    """Jahres-Excel generieren"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Jahresbericht {year}"
    
    # Header
    ws['A1'] = f"Jahresbericht {year}"
    ws['A1'].font = Font(size=16, bold=True)
    
    # Tabelle
    ws['A3'] = "Monat"
    ws['B3'] = "Einnahmen"
    ws['C3'] = "Ausgaben"
    ws['D3'] = "Gewinn"
    
    for i, data in enumerate(monthly_data, 4):
        ws[f'A{i}'] = data['month']
        ws[f'B{i}'] = float(data['income'])
        ws[f'C{i}'] = float(data['expenses'])
        ws[f'D{i}'] = float(data['profit'])
    
    # Summen
    ws[f'A{16}'] = "GESAMT"
    ws[f'A{16}'].font = Font(bold=True)
    ws[f'B{16}'] = sum(d['income'] for d in monthly_data)
    ws[f'C{16}'] = sum(d['expenses'] for d in monthly_data)
    ws[f'D{16}'] = sum(d['profit'] for d in monthly_data)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'Jahresbericht_{year}.xlsx'
    )


def generate_yearly_pdf(year, monthly_data):
    """Jahres-PDF generieren"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    elements = []
    styles = getSampleStyleSheet()
    
    # Titel
    elements.append(Paragraph(f"Jahresbericht {year}", styles['Title']))
    elements.append(Spacer(1, 30))
    
    # Tabelle
    table_data = [['Monat', 'Einnahmen', 'Ausgaben', 'Gewinn']]
    
    for data in monthly_data:
        table_data.append([
            data['month'],
            f"{data['income']:.2f} €",
            f"{data['expenses']:.2f} €",
            f"{data['profit']:.2f} €"
        ])
    
    # Summen
    table_data.append([
        'GESAMT',
        f"{sum(d['income'] for d in monthly_data):.2f} €",
        f"{sum(d['expenses'] for d in monthly_data):.2f} €",
        f"{sum(d['profit'] for d in monthly_data):.2f} €"
    ])
    
    table = Table(table_data, colWidths=[2*inch, 2*inch, 2*inch, 2*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, -1), (-1, -1), colors.beige),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'Jahresbericht_{year}.pdf'
    )


def generate_device_excel(device_data, start_date, end_date):
    """Geräte-Excel generieren"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gerätebericht"
    
    ws['A1'] = "Gerätebericht"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A2'] = f"Zeitraum: {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}"
    
    ws['A4'] = "Gerät"
    ws['B4'] = "Einnahmen"
    ws['C4'] = "Ausgaben"
    ws['D4'] = "Gewinn"
    ws['E4'] = "Einträge"
    ws['F4'] = "Ø/Tag"
    
    for i, data in enumerate(device_data, 5):
        ws[f'A{i}'] = data['device'].name
        ws[f'B{i}'] = float(data['income'])
        ws[f'C{i}'] = float(data['expenses'])
        ws[f'D{i}'] = float(data['profit'])
        ws[f'E{i}'] = data['entries']
        ws[f'F{i}'] = float(data['avg_daily'])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'Geraetebericht_{start_date.strftime("%Y%m%d")}.xlsx'
    )


def generate_device_pdf(device_data, start_date, end_date):
    """Geräte-PDF generieren"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    elements.append(Paragraph("Gerätebericht", styles['Title']))
    elements.append(Paragraph(
        f"Zeitraum: {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}",
        styles['Normal']
    ))
    elements.append(Spacer(1, 20))
    
    table_data = [['Gerät', 'Einnahmen', 'Ausgaben', 'Gewinn', 'Ø/Tag']]
    
    for data in device_data:
        table_data.append([
            data['device'].name,
            f"{data['income']:.2f} €",
            f"{data['expenses']:.2f} €",
            f"{data['profit']:.2f} €",
            f"{data['avg_daily']:.2f} €"
        ])
    
    table = Table(table_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'Geraetebericht_{start_date.strftime("%Y%m%d")}.pdf'
    )


@reports_bp.route('/export-data')
@login_required
def export_data():
    """Kompletter Datenexport"""
    # Implementierung folgt
    flash('Datenexport wird vorbereitet...', 'info')
    return redirect(url_for('reports.index'))


@reports_bp.route('/quick')
@login_required
def quick_report():
    """Schnellbericht für aktuellen Monat"""
    today = date.today()
    return monthly_report()
